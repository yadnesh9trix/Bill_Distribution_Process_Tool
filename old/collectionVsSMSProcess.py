import pandas as pd
import numpy as np
from datetime import datetime,timedelta
import warnings
warnings.filterwarnings('ignore')
from openpyxl import load_workbook

## Define the today's date
today = datetime.today().date()
last = today - timedelta(days=0)
tdy_fmt  =last.strftime("%d%m%Y")


def todaycollection(inppath):
    ## Read collection report YTD
    ytddata_df = pd.read_excel(inppath + f"Paidamount_list_{tdy_fmt}.xlsx", sheet_name='Total')
    ## Replace the property code values in ytd data
    ytddata_df["propertycode"] = ytddata_df["propertycode"].replace("1100900002.10.10", "1100900002.20")
    ytddata_df['propertycode'] = ytddata_df['propertycode'].astype(float)
    ytddata_df['propertycode'] = ytddata_df['propertycode'].apply("{:.02f}".format)
    # ytddata_df = ytddata_df.drop_duplicates('propertycode')
    # ytddata_df[['propertycode', 'pid']] = ytddata_df['propertycode'].str.split(".",expand=True)
    ytddata_df.dropna(subset=['propertycode'], how='all', inplace=True)

    return ytddata_df
#-----------------------------------------------------------------------------------------------------------------------

def Shasti_ytddata9500SMS(inppath,outpth,ytddata_df):
    ## Shasti data 18706
    illegal_fine_plist = pd.read_excel(inppath + "Illegal_Shasti_Plist(18706).xlsx")
    illegal_fine_plist['propertycode'] = illegal_fine_plist['propertycode'].str.replace("/", ".")
    illegal_fine_plist['illegal_flag'] = "illegal_property"
    illegal_fine_plist = illegal_fine_plist[
        ['propertycode', 'otherstotalbal', 'mobileno', 'illegal_flag']]
    filter_mobile_lst = illegal_fine_plist[illegal_fine_plist['mobileno'].notna()]

    ##----------------------------------------------------------------------------------------------------
    ## Merge
    merge_ytd_illegal = ytddata_df.merge(filter_mobile_lst, on='propertycode', how='inner')
    merge_ytd_illegal['receiptdate'] =merge_ytd_illegal['receiptdate'].dt.date.astype(str)

    ## Filter greater than 14th march
    filter_data = merge_ytd_illegal[merge_ytd_illegal['receiptdate'] >= "2023-03-14"]
    filter_ytddata_df = ytddata_df[ytddata_df['receiptdate'] >= "2023-03-14"]

    grp_data = filter_data.groupby(['receiptdate']).agg({'paidamount':'sum','propertycode':'count'}).reset_index()
    grp_data.loc["grand_total"] = grp_data.sum(numeric_only=True)

    ytddata_df_grp = filter_ytddata_df.groupby(['receiptdate']).agg({'paidamount':'sum','propertycode':'count'}).reset_index()
    ytddata_df_grp.loc["grand_total"] = grp_data.sum(numeric_only=True)
    #-------------------------------------------------------------------------------------------------------------------
    writer = pd.ExcelWriter(outpth + "/" + f"ShastiVsCollection(SMSCampign9500){tdy_fmt}.xlsx", engine="xlsxwriter")
    filter_data.to_excel(writer,sheet_name='CollectionVsSMS(SMSCampign9500)', index=False)
    filter_ytddata_df.to_excel(writer,sheet_name='filter_ytddata', index=False)
    grp_data.to_excel(writer,sheet_name='Summary(SMSCampign9500)', index=False)
    ytddata_df_grp.to_excel(writer,sheet_name='SummaryCollection', index=False)
    writer.save()
    writer.close()
    #===================================================================================================================

#-----------------------------------------------------------------------------------------------------------------------
def noarrears_onlycurrentbills51k(inppath,outpth,ytddata_df):
    ## 51k Data
    noarrears_onlycurrentbill_df = pd.read_excel(inppath + "NoRemmaningArrears_OnlyCurrentBill_Plist(75k).xlsx",sheet_name='TotalCurrentBillPlist')
    # noarrears_onlycurrentbill_df = pd.read_excel(inppath + "Curr_Bal-1678-23K.xlsx")
    noarrears_onlycurrentbill_df['propertycode'] = noarrears_onlycurrentbill_df['propertycode'].str.replace("/", ".")
    noarrears_onlycurrentbill_df['noarrears_onlycurrentbill'] = 1
    noarrears_onlycurrentbill_df = noarrears_onlycurrentbill_df[
        ['propertycode', 'current','mobileno','noarrears_onlycurrentbill']]
    filter_onlymobile_lst = noarrears_onlycurrentbill_df[noarrears_onlycurrentbill_df['mobileno'].notna()]
    ##----------------------------------------------------------------------------------------------------
    ##Merge
    filter_onlymobile_lst['propertycode'] = filter_onlymobile_lst['propertycode'].astype(float)
    filter_onlymobile_lst['propertycode'] = filter_onlymobile_lst['propertycode'].apply("{:.02f}".format)

    merge_noarrears_onlycurrentbill_df = ytddata_df.merge(filter_onlymobile_lst, on='propertycode', how='inner')
    merge_noarrears_onlycurrentbill_df['receiptdate'] = merge_noarrears_onlycurrentbill_df['receiptdate'].dt.date.astype(str)
    # merge_ytd_illegal = ytddata_df.merge(noarrears_onlycurrentbill_df, on='propertycode', how='inner')
    # merge_ytd_ = filter_onlymobile_lst.merge(ytddata_df, on='propertycode', how='left')
    # wout_paidSMS = merge_ytd_[merge_ytd_['paid_ty'] != 1]
    # wout_paidSMS.to_excel(outpth + "/" + "WithoutPaidPlist(NoArrearsOnlyCuurentBill_51K).xlsx",index=False)
    # filter_data_1 = merge_ytd_illegal[merge_ytd_illegal['noarrears_onlycurrentbill'] == 1]
    ##----------------------------------------------------------------------------------------------------
    filter_data = merge_noarrears_onlycurrentbill_df[merge_noarrears_onlycurrentbill_df['receiptdate'] >= "2023-03-16"]
    filter_ytddata_df = ytddata_df[ytddata_df['receiptdate'] >= "2023-03-16"]
    ##----------------------------------------------------------------------------------------------------
    grp_data = filter_data.groupby(['receiptdate']).agg({'paidamount':'sum','propertycode':'count'}).reset_index()
    grp_data.loc["grand_total"] = grp_data.sum(numeric_only=True)

    ytddata_df_grp = filter_ytddata_df.groupby(['receiptdate']).agg({'paidamount':'sum','propertycode':'count'}).reset_index()
    ytddata_df_grp.loc["grand_total"] = grp_data.sum(numeric_only=True)
    # pvotable = filter_data.pivot_table(index=['receiptdate'], columns='gatname', values='paidamount')

    #-------------------------------------------------------------------------------------------------------------------
    writer = pd.ExcelWriter(outpth + "/" + "OnlyCntBillVsCollcetion(51K).xlsx", engine="xlsxwriter")
    grp_data.to_excel(writer,sheet_name='SummarySMSCampign51K', index=False)
    ytddata_df_grp.to_excel(writer,sheet_name='SummaryCollection51K', index=False)
    filter_data.to_excel(writer,sheet_name='CollectionVsSMS51K', index=False)
    filter_ytddata_df.to_excel(writer,sheet_name='TotalCollection4M16M', index=False)
    writer.save()
    writer.close()
    # ============================================================================================================================

#-----------------------------------------------------------------------------------------------------------------------
def noarrears_onlycurrentbills75k(inppath,outpth,ytddata_df):
    ##
    noarrears_onlycurrentbill_df = pd.read_excel(inppath + "NoRemmaningArrears_OnlyCurrentBill_Plist(75k).xlsx",sheet_name='TotalCurrentBillPlist')
    noarrears_onlycurrentbill_df['propertycode'] = noarrears_onlycurrentbill_df['propertycode'].str.replace("/", ".")
    # noarrears_onlycurrentbill_df[['propertycode', 'pid']] = noarrears_onlycurrentbill_df['propertycode'].str.split("/", expand=True)
    noarrears_onlycurrentbill_df = noarrears_onlycurrentbill_df.drop_duplicates('propertycode')
    noarrears_onlycurrentbill_df['noarrears_onlycurrentbill'] = 1
    noarrears_onlycurrentbill_df = noarrears_onlycurrentbill_df[
        ['propertycode', 'current','mobileno','noarrears_onlycurrentbill']]
    # filter_onlymobile_lst = noarrears_onlycurrentbill_df[noarrears_onlycurrentbill_df['mobileno'].notna()]

    ##------------------------------------------------------------------------------------------------------------------
    ## Merge
    # noarrears_onlycurrentbill_df['propertycode'] = noarrears_onlycurrentbill_df['propertycode'].astype(float)
    # noarrears_onlycurrentbill_df['propertycode'] = noarrears_onlycurrentbill_df['propertycode'].apply("{:.02f}".format)
    merge_ytd_illegal = ytddata_df.merge(noarrears_onlycurrentbill_df, on='propertycode', how='inner')
    merge_ytd_illegal['receiptdate'] =merge_ytd_illegal['receiptdate'].dt.date.astype(str)
    # merge_ytd_illegal = ytddata_df.merge(noarrears_onlycurrentbill_df, on='propertycode', how='inner')
    # merge_ytd_ = filter_onlymobile_lst.merge(ytddata_df, on='propertycode', how='left')
    # wout_paidSMS = merge_ytd_[merge_ytd_['paid_ty'] != 1]
    # wout_paidSMS.to_excel(outpth + "/" + "WithoutPaidPlist(NoArrearsOnlyCuurentBill_51K).xlsx",index=False)
    # filter_data_1 = merge_ytd_illegal[merge_ytd_illegal['noarrears_onlycurrentbill'] == 1]
    ##------------------------------------------------------------------------------------------------------------------
    filter_data = merge_ytd_illegal[merge_ytd_illegal['receiptdate'] >= "2023-03-16"]
    filter_ytddata_df = ytddata_df[ytddata_df['receiptdate'] >= "2023-03-16"]
    ##------------------------------------------------------------------------------------------------------------------
    grp_data = filter_data.groupby(['receiptdate']).agg({'paidamount':'sum','propertycode':'count'}).reset_index()
    grp_data.loc["grand_total"] = grp_data.sum(numeric_only=True)

    ytddata_df_grp = filter_ytddata_df.groupby(['receiptdate']).agg({'paidamount':'sum','propertycode':'count'}).reset_index()
    ytddata_df_grp.loc["grand_total"] = grp_data.sum(numeric_only=True)
    # pvotable = filter_data.pivot_table(index=['receiptdate'], columns='gatname', values='paidamount')
    #-------------------------------------------------------------------------------------------------------------------
    writer = pd.ExcelWriter(outpth + "/" + "OnlyCntBillVsCollcetion(75K).xlsx", engine="xlsxwriter")
    grp_data.to_excel(writer,sheet_name='SummarySMSCampign75K', index=False)
    ytddata_df_grp.to_excel(writer,sheet_name='SummaryCollection75K', index=False)
    filter_data.to_excel(writer,sheet_name='CollectionVsSMS75', index=False)
    filter_ytddata_df.to_excel(writer,sheet_name='Collection4M16M', index=False)
    writer.save()
    writer.close()
    # ============================================================================================================================

#-----------------------------------------------------------------------------------------------------------------------
def SMSCampign2923_23K(inppath,outpth,ytddata_df):
    noarrears_onlycurrentbill_df = pd.read_excel(inppath + "Curr_Bal-1678-23K.xlsx")
    noarrears_onlycurrentbill_df['noarrears_onlycurrentbill'] = 1
    noarrears_onlycurrentbill_df = noarrears_onlycurrentbill_df[
        ['propertycode', 'mseb_ContactNo','noarrears_onlycurrentbill']]
    filter_onlymobile_lst = noarrears_onlycurrentbill_df[noarrears_onlycurrentbill_df['mseb_ContactNo'].notna()]
    ##----------------------------------------------------------------------------------------------------
    ##Merge
    filter_onlymobile_lst['propertycode'] = filter_onlymobile_lst['propertycode'].astype(float)
    filter_onlymobile_lst['propertycode'] = filter_onlymobile_lst['propertycode'].apply("{:.02f}".format)
    merge_ytd_illegal = ytddata_df.merge(filter_onlymobile_lst, on='propertycode', how='inner')
    merge_ytd_illegal['receiptdate'] =merge_ytd_illegal['receiptdate'].dt.date.astype(str)
    ##----------------------------------------------------------------------------------------------------
    ##----------------------------------------------------------------------------------------------------
    filter_data = merge_ytd_illegal[merge_ytd_illegal['receiptdate'] >= "2023-03-16"]
    filter_ytddata_df = ytddata_df[ytddata_df['receiptdate'] >= "2023-03-16"]
    # filter_data = merge_ytd_illegal[merge_ytd_illegal['receiptdate'] >= "2023-03-01"]
    # filter_ytddata_df = ytddata_df[ytddata_df['receiptdate'] >= "2023-03-01"]

    grp_data = filter_data.groupby(['receiptdate']).agg({'paidamount':'sum','propertycode':'count'}).reset_index()
    grp_data.loc["grand_total"] = grp_data.sum(numeric_only=True)

    ytddata_df_grp = filter_ytddata_df.groupby(['receiptdate']).agg({'paidamount':'sum','propertycode':'count'}).reset_index()
    ytddata_df_grp.loc["grand_total"] = grp_data.sum(numeric_only=True)

    #-------------------------------------------------------------------------------------------------------------------
    writer = pd.ExcelWriter(outpth + "/" + "OnlyCntBillVsCollcetion(2923_23K).xlsx", engine="xlsxwriter")
    grp_data.to_excel(writer,sheet_name='SummarySMSCampign', index=False)
    ytddata_df_grp.to_excel(writer,sheet_name='SummaryCollection', index=False)
    filter_data.to_excel(writer,sheet_name='CollectionVsSMS', index=False)
    filter_ytddata_df.to_excel(writer,sheet_name='Collection4M16M', index=False)
    writer.save()
    writer.close()
    # ============================================================================================================================

#-----------------------------------------------------------------------------------------------------------------------
def SMScampign2476(inppath,outpth,ytddata_df):
    # ## Shasti data 18706
    illegal_fine_plist = pd.read_excel(inppath + "Total_Mapping_Data_Illegal2476.xlsx")
    illegal_fine_plist['propertycode'] = illegal_fine_plist['propertycode'].astype(float)
    illegal_fine_plist['propertycode'] = illegal_fine_plist['propertycode'].apply("{:.02f}".format)

    # illegal_fine_plist['Column1.propertycode'] = illegal_fine_plist['Column1.propertycode'].str.replace("/", ".")
    illegal_fine_plist['2476SMSflag'] = "2476smsproperty"
    illegal_fine_plist = illegal_fine_plist[
        ['propertycode','Column1.mseb_ContactNo','2476SMSflag']]
    # filter_mobile_lst = illegal_fine_plist[illegal_fine_plist['mobileno'].notna()]
    ##----------------------------------------------------------------------------------------------------
    ## Merge
    merge_ytd_illegal = ytddata_df.merge(illegal_fine_plist, on='propertycode', how='inner')
    merge_ytd_illegal['receiptdate'] = merge_ytd_illegal['receiptdate'].dt.date.astype(str)
    filter_data = merge_ytd_illegal[merge_ytd_illegal['receiptdate'] >= "2023-03-15"]
    filter_ytddata_df = ytddata_df[ytddata_df['receiptdate'] >= "2023-03-15"]

    grp_data = filter_data.groupby(['receiptdate']).agg({'paidamount':'sum','propertycode':'count'}).reset_index()
    grp_data.loc["grand_total"] = grp_data.sum(numeric_only=True)

    ytddata_df_grp = filter_ytddata_df.groupby(['receiptdate']).agg({'paidamount':'sum','propertycode':'count'}).reset_index()
    ytddata_df_grp.loc["grand_total"] = grp_data.sum(numeric_only=True)
    #----------------------------------------------------------------------------------------------------------------------------
    writer = pd.ExcelWriter(outpth + "/" + f"ShastiVsCollection(SMSCampign2476){tdy_fmt}.xlsx", engine="xlsxwriter")
    grp_data.to_excel(writer,sheet_name='SummarySMSCampign', index=False)
    ytddata_df_grp.to_excel(writer,sheet_name='YTDfilter', index=False)
    filter_data.to_excel(writer,sheet_name='CollectionVsSMS', index=False)
    filter_ytddata_df.to_excel(writer,sheet_name='filter_ytddata_greaterthn14', index=False)
    writer.save()
    writer.close()
    #===============================================================================================================================

#-----------------------------------------------------------------------------------------------------------------------
## For daily updates graph Outstanding Vs Collection
def ShastiVsCollection18706(inppath,outpth,ytddata_df):
    ## Shasti data 18706
    illegal_fine_plist = pd.read_excel(inppath + "Illegal_Shasti_Plist(18706).xlsx")
    illegal_fine_plist['propertycode'] = illegal_fine_plist['propertycode'].str.replace("/", ".")
    # illegal_fine_plist[['propertycode', 'pid']] = illegal_fine_plist['propertycode'].str.split("/", expand=True)
    # illegal_fine_plist = illegal_fine_plist.drop_duplicates('propertycode')
    illegal_fine_plist['illegal_flag'] = "illegal_property"
    illegal_fine_plist = illegal_fine_plist[
        ['propertycode', 'otherstotalbal', 'mobileno', 'illegal_flag']]
    # filter_mobile_lst = illegal_fine_plist[illegal_fine_plist['mobileno'].notna()]

    ##----------------------------------------------------------------------------------------------------
    ## Merge
    # ytddata_df['PaidFlag'] = 1
    # merge_ytd_illegal = illegal_fine_plist.merge(ytddata_df, on='propertycode', how='left')
    # merge_ytd_illegal.to_excel(outpth + "/" + "shasti_Report(18706)Nitin.xlsx", index=False)

    merge_ytd_illegal = ytddata_df.merge(illegal_fine_plist, on='propertycode', how='inner')
    merge_ytd_illegal['receiptdate'] = merge_ytd_illegal['receiptdate'].dt.date.astype(str)
    filter_ytddata_df = merge_ytd_illegal[merge_ytd_illegal['receiptdate'] >= "2023-03-01"]

    #---------------------------------------------------------------------------------------------------------------------------------
    writer = pd.ExcelWriter(outpth + "/" + f"ShastiVsCollection(18706_01March){tdy_fmt}.xlsx", engine="xlsxwriter")
    merge_ytd_illegal.to_excel(writer,sheet_name=f'ShastiMatch_ytd{tdy_fmt}', index=False)
    filter_ytddata_df.to_excel(writer,sheet_name='ShastiVsCollection01March', index=False)
    writer.save()
    writer.close()
    #=========================================================

#-----------------------------------------------------------------------------------------------------------------------
def ArrearsCrntBillsPending_WoutShastiPlist_1lac10K(inppath,outpth,ytddata_df):
    arrears_onlycurrentbill_df = pd.read_excel(inppath + "Arrears&CurrentBillsPending_Plist(WoutShastiPlist).xlsx")
    arrears_onlycurrentbill_df['propertycode'] = arrears_onlycurrentbill_df['propertycode'].str.replace("/", ".")
    arrears_onlycurrentbill_df['Flag'] = 1
    arrears_onlycurrentbill_df = arrears_onlycurrentbill_df[
        ['propertycode', 'current','total','mobile_no','Flag']]
    # filter_onlymobile_lst = noarrears_onlycurrentbill_df[noarrears_onlycurrentbill_df['mobileno'].notna()]
    ##----------------------------------------------------------------------------------------------------
    ##Merge
    # arrears_onlycurrentbill_df['propertycode'] = arrears_onlycurrentbill_df['propertycode'].astype(float)
    # arrears_onlycurrentbill_df['propertycode'] = arrears_onlycurrentbill_df['propertycode'].apply("{:.02f}".format)
    merge_arrears_onlycurrentbill = ytddata_df.merge(arrears_onlycurrentbill_df, on='propertycode', how='inner')
    merge_arrears_onlycurrentbill['receiptdate'] =merge_arrears_onlycurrentbill['receiptdate'].dt.date.astype(str)
    ##----------------------------------------------------------------------------------------------------
    filter_data = merge_arrears_onlycurrentbill[merge_arrears_onlycurrentbill['receiptdate'] >= "2023-03-22"]
    filter_ytddata_df = ytddata_df[ytddata_df['receiptdate'] >= "2023-03-22"]
    ##----------------------------------------------------------------------------------------------------
    grp_data = filter_data.groupby(['receiptdate']).agg({'paidamount':'sum','propertycode':'count'}).reset_index()
    grp_data.loc["grand_total"] = grp_data.sum(numeric_only=True)

    ytddata_df_grp = filter_ytddata_df.groupby(['receiptdate']).agg({'paidamount':'sum','propertycode':'count'}).reset_index()
    ytddata_df_grp.loc["grand_total"] = grp_data.sum(numeric_only=True)

    #-------------------------------------------------------------------------------------------------------------------
    writer = pd.ExcelWriter(outpth + "/" + "ArrearsCntBillPndingVsCollcetion(1lac10K).xlsx", engine="xlsxwriter")
    grp_data.to_excel(writer,sheet_name='SummarySMSCampign(1lac10K)', index=False)
    ytddata_df_grp.to_excel(writer,sheet_name='SummaryTotalCollection', index=False)
    filter_data.to_excel(writer,sheet_name='CollectionVsSMS', index=False)
    filter_ytddata_df.to_excel(writer,sheet_name='TotalCollectionList', index=False)
    writer.save()
    writer.close()
    # ============================================================================================================================

#-----------------------------------------------------------------------------------------------------------------------
def total5yearscollcetion(inppath,outpth,ytddata_df):

    def get_sheetnames_xlsx(filepath):
        wb = load_workbook(filepath, read_only=True, keep_links=False)
        return wb.sheetnames
    infile = inppath + "Paid_amountList(Last4Year).xlsx"
    sheetname = get_sheetnames_xlsx(infile)
    lst = []
    for i in sheetname:
        totalPaidAmount = pd.read_excel(inppath + "Paid_amountList(Last4Year).xlsx", sheet_name=i)
        lst.append(totalPaidAmount)
    ddfff = pd.DataFrame(pd.concat(lst))

    newwdd = ddfff.append(ytddata_df)

    newwdd['receiptdate'] = pd.to_datetime(newwdd['receiptdate'], errors='coerce', format='%Y-%m-%d')
    newwdd['fin_year_r'] = newwdd['receiptdate'].dt.year
    newwdd['fin_month_r'] = newwdd['receiptdate'].dt.month

    newwdd['fin_year_r'] = np.where(newwdd['fin_month_r'] < 4,
                                    newwdd['fin_year_r'] - 1,
                                    newwdd['fin_year_r'])
    newwdd['fin_month_r'] = np.where(newwdd['fin_month_r'] < 4,
                                     newwdd['fin_month_r'] + 9,
                                     newwdd['fin_month_r'] - 3)

    finYr2018_19 = newwdd[(newwdd['fin_year_r']>=2018) & (newwdd['fin_year_r']<2019)]
    finYr2019_20 = newwdd[(newwdd['fin_year_r']>=2019) & (newwdd['fin_year_r']<2020)]
    finYr2020_21 = newwdd[(newwdd['fin_year_r']>=2020) & (newwdd['fin_year_r']<2021)]
    finYr2021_22 = newwdd[(newwdd['fin_year_r']>=2021) & (newwdd['fin_year_r']<2022)]
    finYr2022_23 = newwdd[(newwdd['fin_year_r']>=2022) & (newwdd['fin_year_r']<2023)]

    aaa = finYr2018_19.append(finYr2019_20)
    aaa1 = aaa.append(finYr2020_21)
    aaa2  = aaa1.append(finYr2021_22)
    aaa12 = aaa2.append(finYr2022_23)

    grrp = aaa12.groupby(['fin_year_r', 'modeofpayment']).agg(
        {'propertycode': 'count', 'paidamount': 'sum'}).reset_index()
    # grrp_df = grrp.rename(columns={'propertycode':'No.Of Property'})
    pvotable = grrp.pivot_table(index=['modeofpayment'], columns=['fin_year_r'], values='paidamount')
    pvotable.loc['Grand total'] = pvotable.sum(numeric_only=True)

    pvotable = pvotable.fillna(0).reset_index()
    pvotable.to_excel(outpth + "/" + "SummaryLast5YearsModeOfPay.xlsx", index=False)
